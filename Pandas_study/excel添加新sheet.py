# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import os

"""
#enter code here
# dataframe: 需要写入excel的数据
# outfile：输出的文件地址
# name: sheet_name的文件名称
def excelAddSheet(dataframe, outfile, name):
    writer = pd.ExcelWriter(outfile, enginge='openpyxl')
    if os.path.exists(outfile) != True:
        dataframe.to_excel(writer, name, index=None)
    else:
        book = load_workbook(writer.path)
        writer.book = book
        dataframe.to_excel(excel_writer=writer, sheet_name = name, index=None)
    writer.save()
    writer.close()
"""
students = pd.read_excel('./excel/Students-excelwriter.xlsx')
print(students.head())

outfile = './excel/Students-excelwriter123.xlsx'
writer = pd.ExcelWriter(outfile, engine='openpyxl')
print(writer.path)

# 不存在进行保存，上面例子中使用writer的不能进行保存
if os.path.exists(outfile) != True:
    students.to_excel(outfile, sheet_name='abc', index=None)
    print('save one new excel.')
else:
    # 必须加load_workbook 和writer.book 这两行
    book = load_workbook(writer.path)
    writer.book = book
    students.to_excel(excel_writer=writer, sheet_name='hello', index=None)
    writer.save()
    writer.close()
    print('add a new sheet in the excel sucess.')
