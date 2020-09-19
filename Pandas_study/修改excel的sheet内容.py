# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import os

"""
# 读取后drop
data = pd.read_excel("data.xlsx", sheet_name=sheet_name)
mydata = data.drop([0], axis=0)
# 保存新的数据
book = load_workbook('data.xlsx')
writer = pd.ExcelWriter('data.xlsx',engine='openpyxl')
writer.book = book
# 清除原来的数据
idx = book.sheetnames.index('mysheet')
book.remove(book.worksheets[idx])
book.create_sheet('mysheet', idx)

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
mydata.to_excel(writer, "mysheet")
writer.save()
"""
# students = pd.read_excel('./excel/Students-excelwriter.xlsx', sheet_name=[0,1])
mydata = pd.read_excel('./excel/Students-excelwriter123.xlsx')
# print(students)
mydata.drop(index=mydata[7:20].index, inplace=True)
book = load_workbook('./excel/Students-excelwriter123.xlsx')
writer = pd.ExcelWriter('./excel/Students-excelwriter123.xlsx', engine='openpyxl')
writer.book = book


# 不删除原有的sheet的话，最后修改只会让原来的值进行覆盖已有的单元格，
# 同sheet中其它数据并不会被清除，因此需要删除在通过索引创建一个新的填写
# 找到自己要修改的sheet索引，使用索引还有一个目的就是插入sheet
# sheet名为 hello的索引
idx = book.sheetnames.index('hello')    # 0
# print(book.worksheets[0])       # <Worksheet "hello">
# print(book['hello'])        # <Worksheet "hello">

# 删除那个不要的sheet
book.remove(book.worksheets[idx])
book.create_sheet('hello', idx)

# 这个也是必不可少
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
mydata.to_excel(writer, "hello")

# 不保存就可以重复调试了，一旦保存重复执行会出现数据问题
# writer.save()

print(mydata)
print(writer.sheets)
