# -*- coding: utf-8 -*-

### PermissionError: [Errno 13] Permission denied: 'Student1.xlsx'
# 报权限问题，是因为外面打开了excel表格，关闭重新执行即可
"""
<1>openpyxl的三种最重要的数据结构
NULL空值：对应于python中的None，表示这个cell里面没有数据。
numberic： 数字型，统一按照浮点数来进行处理。对应于python中的float。
string： 字符串型，对应于python中的unicode。
<2>Excel 文件的三个对象
workbook： 工作簿，一个excel工作簿包含多个sheet。
sheet：工作表，一个workbook有多个，表名识别，如“sheet1”,“sheet2”等。
cell： 单元格，存储数据对象
<3>导入
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors, fills
from openpyxl.formatting.rule import ColorScale

# <4> 加载workbook
# 加载workbook时注意openpyxl只支持xlsx格式，老版的xls格式需要其他方法去加载。
wb = load_workbook('Student1.xlsx')

# <5> 新建worksheet
# 新建工作表用 函数create_sheet()

# ws1 = wb.create_sheet() # 默认插在最后
# ws2 = wb.create_sheet(0)    # 插在开头建表后默认名按顺序，如：sheet1，sheet2...
# ws = wb['Sheet2']
# ws.title = "New Title"  # 修改表名
# print(ws)
# 创建时可以直接给ws2 = wb.create_sheet(title='New Title')

# <6> 打开worksheet

# 1) 通过名字
ws = wb['Sheet1']
print(ws)
# 2) 不知道名字使用index
sheet_name_list = wb.sheetnames
ws = wb[sheet_name_list[1]]
# print(sheet_name_list[0])
print(ws)

# 3) 获取活动表表名
# print(wb.get_active_sheet().title)    # 不推荐函数，不要使用
print(wb.active)
print('###title: ' ,wb.active.title)

# <7> 编辑单元格
# 1) 读写单个单元格
ws = wb[sheet_name_list[1]]
print('ws: ', ws)
c = ws['A4']    # read 等同于 c = ws.cell(row, column)
print(c)
print(ws.cell(1, 3).value)  # 输出第一行第三列的按个值
# ws['A4'] = 4 # write
# 行列读写
# 给单元格赋值
print("hereherehereherehereherehereherehereherehereherehere")
ws.cell(row=4, column=2).value='test'
print(ws.cell(row=4, column=2).value)
# 2) 读写多个单元格
cell_range = ws['A1':'C2']  # #读取A1到C2单元格数据
print(cell_range)   # 值为嵌套元组，因此需要双层循环求值
for cell in cell_range:
    for i in cell:print(i.value)

# <8> 获取最大行和最大列
print('max_row', ws.max_row)
print('max_column', ws.max_column)

# <9> 获取行和列的值
print("######")
for row in ws.rows:
    for i in row:print(i.value, end=";")
for column in ws.columns:
    for i in column:print(i.value, end=";")

wb.save('Student1.xlsx')
"""
上面的代码就可以获得所有单元格的数据。如果要获得某行的数据呢？因sheet.rows是生成器类型，不能使用索引，但是转换成list之后便可使用索引，如list(sheet.rows)[2]这样就获取到第三行的tuple对象。
for cell in list(sheet.rows)[2]:
    print(cell.value)
也可以通过iter_rows的参数min_row, max_row指定并获取指定行内容，如下所示，获取第二行表格的内容:

for row in ws.iter_rows(min_row=2, max_row=2,):
    line = [cell.value for cell in row]
    print(line)

此外可使用方法ws.iter_rows()逐行读取数据
ws.iter_rows(range_string=None, row_offset=0, column_offset=0): 
range-string(string)-单元格的范围：例如('A1:C4') row_offset-添加行 column_offset-添加列 
# 返回一个生成器, 注意取值时要用value,例如：
wb = load_workbook('testdata.xlsx')
ws = wb.active
ws.title = 'testdata.xlsx'
for row in ws.iter_rows('A1:C2'):
    for cell in row:
        print(cell.value)
"""

#########################################################################
print("\nxxxxxxxx\n")
from openpyxl import load_workbook

"""
<10>获取任意区间的值
方法1、使用range函数
可以使用range函数，下面的写法，获得了以A1为左上角，B3为右下角矩形区域的所有单元格。注意range从1开始的，因为在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值。
"""
wb = load_workbook('Student2.xlsx')
ws = wb['Sheet1']
print(ws)
for i in range(1, 4):
    for j in range(1, 3):
        print(ws.cell(row=i, column=j))

# 方法1、
# for row_cell in ws['A1':'B3']:
#     for cell in row_cell:
#         print(cell)

# <11> 根据字母获得列号，根据列号返回字母
from openpyxl.utils import get_column_letter, column_index_from_string

print("# <11> 根据字母获得列号，根据列号返回字母")
# 根据列的数字返回字母
print(get_column_letter(2)) # B
print(column_index_from_string('C'))    # 3

# 写单元格
print("# 写单元格")
# 直接给单元格赋值,加不加value都可以
# ws['A1'] = 'good'
ws['A1'].value = 'good'
print('A1:', ws['A1'])
print(ws['A1'].value)
# 利用excel自身公式写入值
print("# 利用excel自身公式写入值")
# 利用公式在B2处写入平均值
ws['B2'] = "=AVERAGE(B3:B6)"
msg = """
    注意：如果是读取，需要加上data_only=True，这样读到B2返回的才是数字，
    如：wb = load_workbook('Student2.xlsx', data_only=True)；
    如果不加这个参数，将是返回这个公式本身 "=AVERAGE(B3:B6)"
"""
print(msg)

# <13> 使用append函数写数据
print("append函数可以一次添加多行数据，从第一行空白行开始（下面都是空白行）写入")
# ws.append(iterable)
"""
 #添加一行到当前sheet的最底部（即逐行追加从第一行开始） iterable必须是list,tuple,dict,range,generator类型的。 1,如果是list,将list从头到尾顺序添加。 2，如果是dict,按照相应的键添加相应的键值。
 ws.append(['This is A1', 'This is B1', 'This is C1'])
 ws.append({'A' : 'This is A1', 'C' : 'This is C1'})
 ws.append({1 : 'This is A1', 3 : 'This is C1'})
# 添加一行如下：
"""
row = [1 ,2, 3, 4, 5]
ws.append(row)

wb.save('Student2.xlsx')

# ws.iter_rows的使用
wb = load_workbook('Student2.xlsx')
ws = wb['New Title']
for row in ws.iter_rows(min_row=1, max_row=1, ):
    title = [cell.value for cell in row]
print(title)

for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=3):
    title = [cell.value for cell in row]
print(title)
