# -*- coding: utf-8 -*-
# 网易云课堂 Python ABC
#  https://study.163.com/course/courseMain.htm?courseId=1004106037

import openpyxl
from openpyxl.styles import Font    # Color 用16进制表示，对新手不友好
from openpyxl.styles import colors  # 此colors使用英文单词表示

# 创建一个工作簿
wb = openpyxl.Workbook()

# Font
ws = wb.active
ws.title = 'Font'

bold_red_font = Font(name='Times New Roman', bold=True, color=colors.RED)
ws['A1'].font = bold_red_font
ws['A1'] = 'Bold Red Times New Roman'

# default 11pt, Calibri
italic24_font = Font(size=24, italic=True)
ws['B3'].font = italic24_font
ws['B3'] = '24pt Italic'

# Formulas
ws = wb.create_sheet('Formula')
ws['A1'] = 200
ws['A2'] = 300
ws['A3'] = '=SUM(A1:A2)'
# 整行添加数据
ws.append(['xxx', '666'])

# setting row height and column width
ws = wb.create_sheet('dimensions')
ws['A1'] = 'Tall row'
ws.row_dimensions[1].height = 70
ws['B2'] = 'Wide column'
ws.column_dimensions['B'].width = 20

# merging cells
ws = wb.create_sheet('merged')
ws.merge_cells('A1:D3')
ws['A1'] = 'Twelve cells merged together'
ws.merge_cells('C5:D5')
ws['C5'] = 'Two merged cells'

print('C5: ', ws['C5'].value)

# unmerging cells
# Call to deprecated function get_sheet_by_name (Use wb[sheetname])
# ws = wb.copy_worksheet(wb.get_sheet_by_name('merged'))
ws = wb.copy_worksheet(wb['merged'])
ws.title = 'unmerged'
ws.unmerge_cells('A1:D3')
ws.unmerge_cells('C5:D5')

wb.save('excel_styles.xlsx')
