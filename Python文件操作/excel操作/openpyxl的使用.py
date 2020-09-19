# -*- coding: utf-8 -*-

import openpyxl
import os, sys


"""
# 1.1 Workbook提供的部分常用属性如下：
active：获取当前活跃的Worksheet
worksheets：以列表的形式返回所有的Worksheet(表格)
read_only：判断是否以read_only模式打开Excel文档
write_only:判断是否以write_only模式打开Excel文档
encoding：获取文档的字符集编码
properties：获取文档的元数据，如标题，创建者，创建日期等
sheetnames：获取工作簿中的表（列表）
"""
workbook_path = os.path.join("Student.xlsx")
wb = openpyxl.load_workbook(workbook_path)
ws = wb.active
print(ws)   # <Worksheet "Sheet1">  光标所在的的Sheet
ws_list = wb.worksheets
print(ws_list)  # [<Worksheet "Sheet1">, <Worksheet "Sheet2">, <Worksheet "Sheet3">]
print(wb.read_only) # False
print(wb.encoding)  # utf-8
print(wb.write_only)    # False
# print(wb.properties)
# <openpyxl.packaging.core.DocumentProperties object>
# Parameters:
# creator='Administrator', title=None, description=None, subject=None, identifier=None, language=None, created=datetime.datetime(2019, 11, 30, 12, 57),
#  modified=datetime.datetime(2019, 12, 7, 6, 8, 14), lastModifiedBy='Administrator', category=None, contentStatus=None, version=None, revision=None,
# keywords=None, lastPrinted=None

"""
1.2  Workbook提供的部分常用方法如下：
get_sheet_names：获取所有表格的名称(新版已经不建议使用，通过Workbook的sheetnames属性即可获取)
get_sheet_by_name：通过表格名称获取Worksheet对象(新版也不建议使用，通过Worksheet['表名']获取)
get_active_sheet：获取活跃的表格(新版建议通过active属性获取)
remove_sheet：删除一个表格
create_sheet：创建一个空的表格
copy_worksheet：在Workbook内拷贝表格
"""

workbook_path = os.path.join("Student.xlsx")
wb = openpyxl.load_workbook(workbook_path)
# DeprecationWarning: Call to deprecated function get_sheet_names
# 调用不推荐的函数告警，虽然能打印，但告警更换sheetnames方式
# sheet_name_list = wb.get_sheet_names()
# print(sheet_name_list)
sheet_name_list = wb.sheetnames
print(sheet_name_list)  # ['Sheet1', 'Sheet2', 'Sheet3']
# ws1 = wb.get_sheet_by_name('Sheet1')
ws1 = wb['Sheet1']
print(ws1)  # <Worksheet "Sheet1">
ws2 = wb.active
print(ws2)  # <Worksheet "Sheet1">
ws3 = wb.create_sheet('text')   # 创建一个表，默认是末尾；有则默认创建一个同名加数字的名称
print('ws3:', ws3)
print(wb.sheetnames)   # ['Sheet1', 'Sheet2', 'Sheet3', 'text']
wb.remove(ws3)
# 注意删除的格式
print(wb.sheetnames)    # ['Sheet1', 'Sheet2', 'Sheet3']
# wb.copy_worksheet(from_worksheet=ws1)
print(wb.index(wb['text1']))
wb.save('Student.xlsx')
